import xlrd
import xlsxwriter
import os
import openpyxl as openpyxl

#文件扫描
def fileScanner(path):
    for dirPath,dirNames,fileNames in os.walk(path):
        #print(dirPath,dirNames,fileNames)
        return fileNames

#文件生成
def fileCreater(path,filenames):
    for filename in filenames:
        filePath=path+'\\'+filename
        print(filePath)
        try:
            desWorkBook=xlsxwriter.Workbook(filePath)
            desWorkBook.add_worksheet(filename)
            desWorkBook.close()
        except Exception as e:
            print('创建错误：',e)

#内容复制
def copyData(sourceWorkBookPath,desWorkBookPath):
    #获取源数据
    sourceWorkBook=xlrd.open_workbook(sourceWorkBookPath)
    sourceSheet=sourceWorkBook.sheets()[1]
    #激活生成的sheet
    desWorkBook=openpyxl.load_workbook(desWorkBookPath)
    desSheet=desWorkBook.active
    #获取源工作表行数
    sourceSheetRows=sourceSheet.nrows
    data=[]
    for row in range(sourceSheetRows):
        data.append(sourceSheet.row_values(row))
    dataRows=len(data)
    dataCols=len(data[0])
    for row in range(dataRows):
        for cols in range(dataCols):
            print(row+1,cols+1,data[row][cols])
            desSheet.cell(row=row+1,column=cols+1,value=data[row][cols])
    desWorkBook.save(desWorkBookPath)
    desWorkBook.close()

if __name__ == '__main__':
    # 文件操作目录
    # 生成的文件目录
    desWorkBookDirPath = r'C:\Users\97759\Desktop\project\20220111\results'
    # 源文件目录
    sourceWorkBookDirPath = r'C:\Users\97759\Desktop\project\20220111'

    filenames=fileScanner(sourceWorkBookDirPath)
    #print(filenames)
    fileCreater(desWorkBookDirPath,filenames)
    for filename in filenames:
        sourceWorkBookPath=sourceWorkBookDirPath+'\\'+filename
        desWorkBookPath=desWorkBookDirPath+'\\'+filename
        copyData(sourceWorkBookPath,desWorkBookPath)
